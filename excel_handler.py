"""
src/excel_handler.py
Baca, duplikat, dan update file Excel test case.
Semua operasi write dilakukan ke file testcase_result.xlsx di folder result,
bukan ke file master testcase.xlsx.
"""

import logging
import shutil
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

logger = logging.getLogger(__name__)

# ── Konstanta warna status ─────────────────────────────────────────────────────
COLOR_PASS   = "D6F5E3"  # Hijau muda
COLOR_FAIL   = "FFE0E0"  # Merah muda
COLOR_ERROR  = "FFF3CD"  # Kuning muda
FONT_PASS    = "1A7A3C"  # Hijau tua
FONT_FAIL    = "CC0000"  # Merah tua
FONT_ERROR   = "B8860B"  # Kuning tua

# Kolom wajib di file test case
REQUIRED_COLUMNS = ["test_case_id", "policy_number", "product_code", "status"]


def read_testcase(testcase_path: str) -> pd.DataFrame:
    """
    Baca file Excel test case dan validasi kolom wajib.

    Args:
        testcase_path: Path ke file testcase.xlsx

    Returns:
        DataFrame test case

    Raises:
        FileNotFoundError: Jika file tidak ditemukan
        ValueError: Jika kolom wajib tidak ada
    """
    path = Path(testcase_path)
    if not path.exists():
        raise FileNotFoundError(f"File test case tidak ditemukan: {testcase_path}")

    df = pd.read_excel(testcase_path, dtype=str)
    df.columns = df.columns.str.strip()

    missing = [col for col in REQUIRED_COLUMNS if col not in df.columns]
    if missing:
        raise ValueError(
            f"File test case tidak lengkap.\n"
            f"Kolom yang hilang: {missing}\n"
            f"Kolom wajib: {REQUIRED_COLUMNS}"
        )

    logger.info(f"Test case dibaca: {len(df)} policy ditemukan.")
    return df


def duplicate_testcase(testcase_path: str, result_folder: str) -> str:
    """
    Duplikat file test case ke folder result dengan nama testcase_result.xlsx.

    Args:
        testcase_path: Path file test case master
        result_folder: Path folder result (sudah harus ada)

    Returns:
        Path file hasil duplikat

    Raises:
        FileNotFoundError: Jika file master tidak ditemukan
    """
    src = Path(testcase_path)
    if not src.exists():
        raise FileNotFoundError(f"File test case tidak ditemukan: {testcase_path}")

    dst = Path(result_folder) / "testcase_result.xlsx"
    shutil.copy2(src, dst)
    logger.info(f"Test case diduplikat ke: {dst}")
    return str(dst)


def fill_db_data(result_path: str, policy_number: str, db_data: dict) -> None:
    """
    Tulis data dari DB ke kolom-kolom detail di baris yang sesuai policy_number.
    Hanya kolom yang ada di db_data yang diupdate.

    Args:
        result_path: Path file testcase_result.xlsx
        policy_number: Nomor polis yang dicari
        db_data: dict {nama_kolom: nilai} dari hasil query DB
    """
    if not db_data:
        logger.warning(f"DB data kosong untuk policy: {policy_number}, skip fill.")
        return

    backup = _backup_file(result_path)
    try:
        wb = load_workbook(result_path)
        ws = wb.active

        headers = {cell.value: cell.column for cell in ws[1] if cell.value}
        policy_col = headers.get("policy_number")
        if not policy_col:
            raise ValueError("Kolom 'policy_number' tidak ditemukan di file result.")

        for row in ws.iter_rows(min_row=2):
            if str(row[policy_col - 1].value).strip() == str(policy_number).strip():
                for col_name, value in db_data.items():
                    if col_name in headers:
                        ws.cell(row=row[0].row, column=headers[col_name], value=value)
                break

        wb.save(result_path)
        _remove_backup(backup)
        logger.info(f"DB data ditulis ke result untuk policy: {policy_number}")

    except Exception as e:
        _restore_backup(backup, result_path)
        raise RuntimeError(f"Gagal menulis DB data untuk policy {policy_number}: {e}")


def fill_transaction_data(
    result_path: str,
    policy_number: str,
    transactions: list[dict],
) -> None:
    """
    Tulis data transaksi dari DB ke sheet "DB Transactions" di testcase_result.xlsx.
    Sheet dibuat otomatis jika belum ada.
    Baris lama untuk policy yang sama akan dihapus dulu sebelum ditulis ulang.

    Args:
        result_path: Path file testcase_result.xlsx
        policy_number: Nomor polis
        transactions: list of dict hasil query DB transaksi
    """
    if not transactions:
        logger.warning(f"Data transaksi kosong untuk policy: {policy_number}, skip.")
        return

    backup = _backup_file(result_path)
    try:
        wb = load_workbook(result_path)

        # Buat sheet jika belum ada
        sheet_name = "DB Transactions"
        if sheet_name not in wb.sheetnames:
            ws_tx = wb.create_sheet(sheet_name)
            _build_transaction_sheet_header(ws_tx, transactions[0].keys())
        else:
            ws_tx = wb[sheet_name]

        # Hapus baris lama untuk policy ini (hindari duplikat jika run ulang)
        rows_to_delete = []
        for row in ws_tx.iter_rows(min_row=2):
            if row[0].value and str(row[0].value).strip() == str(policy_number).strip():
                rows_to_delete.append(row[0].row)
        for row_num in reversed(rows_to_delete):
            ws_tx.delete_rows(row_num)

        # Tulis data transaksi baru
        from openpyxl.styles import Border, Side
        border_side = Side(style="thin", color="CCCCCC")
        thin_border = Border(
            left=border_side, right=border_side,
            top=border_side,  bottom=border_side
        )
        alt_fill = PatternFill("solid", start_color="EEF4FF", end_color="EEF4FF")
        wht_fill = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")

        next_row = ws_tx.max_row + 1
        for i, tx in enumerate(transactions):
            bg = alt_fill if i % 2 == 0 else wht_fill
            # Kolom pertama selalu policy_number
            c = ws_tx.cell(row=next_row, column=1, value=policy_number)
            c.font      = Font(name="Arial", size=9)
            c.fill      = bg
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = thin_border

            for col_idx, (key, val) in enumerate(tx.items(), start=2):
                cell = ws_tx.cell(row=next_row, column=col_idx, value=val)
                cell.font      = Font(name="Arial", size=9)
                cell.fill      = bg
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border    = thin_border
            ws_tx.row_dimensions[next_row].height = 17
            next_row += 1

        wb.save(result_path)
        _remove_backup(backup)
        logger.info(
            f"Data transaksi ditulis ke sheet '{sheet_name}': "
            f"{len(transactions)} baris untuk policy {policy_number}"
        )

    except Exception as e:
        _restore_backup(backup, result_path)
        raise RuntimeError(
            f"Gagal menulis data transaksi untuk policy {policy_number}: {e}"
        )


def _build_transaction_sheet_header(ws, column_keys) -> None:
    """Buat header row untuk sheet DB Transactions."""
    from openpyxl.styles import Border, Side
    border_side  = Side(style="thin", color="CCCCCC")
    thin_border  = Border(
        left=border_side, right=border_side,
        top=border_side,  bottom=border_side
    )
    header_fill  = PatternFill("solid", start_color="003366", end_color="003366")
    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers    = ["policy_number"] + list(column_keys)
    col_widths = {
        "policy_number":         18,
        "transaction_date":      22,
        "transaction_type":      25,
        "price_date":            22,
        "unit_amount":           14,
        "unit_price":            14,
        "transaction_value":     22,
        # fallback lebar default
    }

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = thin_border

        # Set lebar kolom
        from openpyxl.utils import get_column_letter
        width = col_widths.get(header, 20)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"


def update_row_status(
    result_path: str,
    test_case_id: str,
    status: str,
    mismatch_fields: list[str] | None = None,
    notes: str | None = None,
) -> None:
    """
    Update kolom status, mismatch_fields, dan notes di baris yang sesuai test_case_id.
    Warna sel status disesuaikan: PASS=hijau, FAIL=merah, ERROR=kuning.

    Args:
        result_path: Path file testcase_result.xlsx
        test_case_id: ID test case yang akan diupdate
        status: "PASS", "FAIL", atau "ERROR"
        mismatch_fields: list nama field yang tidak cocok (opsional)
        notes: Catatan tambahan atau pesan error (opsional)
    """
    backup = _backup_file(result_path)
    try:
        wb = load_workbook(result_path)
        ws = wb.active

        headers   = {cell.value: cell.column for cell in ws[1] if cell.value}
        id_col    = headers.get("test_case_id")
        stat_col  = headers.get("status")
        mis_col   = headers.get("mismatch_fields")
        note_col  = headers.get("notes")

        if not id_col or not stat_col:
            raise ValueError("Kolom 'test_case_id' atau 'status' tidak ditemukan.")

        bg_color, font_color = _get_status_colors(status)

        for row in ws.iter_rows(min_row=2):
            if str(row[id_col - 1].value).strip() == str(test_case_id).strip():
                # Update status
                stat_cell = ws.cell(row=row[0].row, column=stat_col)
                stat_cell.value     = status
                stat_cell.fill      = PatternFill("solid", start_color=bg_color,   end_color=bg_color)
                stat_cell.font      = Font(bold=True, color=font_color)
                stat_cell.alignment = Alignment(horizontal="center")

                # Update mismatch_fields
                if mis_col and mismatch_fields is not None:
                    ws.cell(row=row[0].row, column=mis_col,  value=", ".join(mismatch_fields))

                # Update notes
                if note_col and notes is not None:
                    ws.cell(row=row[0].row, column=note_col, value=notes)

                break

        wb.save(result_path)
        _remove_backup(backup)
        logger.info(f"Status diupdate: {test_case_id} → {status}")

    except Exception as e:
        _restore_backup(backup, result_path)
        raise RuntimeError(f"Gagal update status untuk {test_case_id}: {e}")


# ── Private Helpers ───────────────────────────────────────────────────────────

def _get_status_colors(status: str) -> tuple[str, str]:
    """Return (background_color, font_color) berdasarkan status."""
    mapping = {
        "PASS":  (COLOR_PASS,  FONT_PASS),
        "FAIL":  (COLOR_FAIL,  FONT_FAIL),
        "ERROR": (COLOR_ERROR, FONT_ERROR),
    }
    return mapping.get(status.upper(), ("FFFFFF", "000000"))


def _backup_file(path: str) -> str:
    """Buat backup sementara sebelum write. Return path backup."""
    backup = path + ".bak"
    shutil.copy2(path, backup)
    return backup


def _remove_backup(backup_path: str) -> None:
    """Hapus file backup setelah write sukses."""
    p = Path(backup_path)
    if p.exists():
        p.unlink()


def _restore_backup(backup_path: str, original_path: str) -> None:
    """Restore file dari backup jika write gagal."""
    p = Path(backup_path)
    if p.exists():
        shutil.copy2(backup_path, original_path)
        p.unlink()
        logger.warning(f"File di-restore dari backup: {original_path}")
